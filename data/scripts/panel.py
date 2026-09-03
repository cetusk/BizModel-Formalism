import openpyxl, glob, re, statistics as st
def load(fn):
    wb=openpyxl.load_workbook(fn, data_only=True)
    ws=wb[wb.sheetnames[0]]
    cols={str(ws.cell(6,c).value):c for c in range(9,ws.max_column+1) if ws.cell(6,c).value}
    rows={str(ws.cell(r,9).value):r for r in range(7,ws.max_row+1) if ws.cell(r,9).value}
    def g(item,col):
        if item not in rows or col not in cols: return None
        v=ws.cell(rows[item],cols[col]).value
        return v if isinstance(v,(int,float)) else None
    return g
files={}
for fn in glob.glob('../raw/chusho/extracted/**/*.xlsx', recursive=True):
    m=re.search(r'r(\d\d)_(\d\d)_1_', fn)
    if m: files[(int(m.group(1)), int(m.group(2)))]=fn
years=sorted({k[0] for k in files})
fy={2:2019,3:2020,4:2021,5:2022,6:2023,7:2024}
tiers=[("合計_法人企業_5人以下","法人5人以下"),("合計_法人企業_6～20人","法人6-20"),
       ("合計_法人企業_21～50人","法人21-50"),("合計_法人企業_51人以上","法人51人~"),
       ("合計_個人企業","個人企業")]
print("=== 営業利益率の推移（決算年度）===")
print(f"{'区分':<12}" + "".join(f"{fy[y]:>8}" for y in years) + f"{'平均':>8}{'SD':>7}")
print("-"*72)
data={}
for k,l in tiers:
    row=f"{l:<12}"; vals=[]
    for y in years:
        g=load(files[(y,3)]); s=g("売上高",k); op=g("営業利益",k)
        if s and op is not None:
            v=op/s*100; vals.append(v); row+=f"{v:>8.1f}"
        else: row+=f"{'--':>8}"
    data[l]=vals
    row+=f"{st.mean(vals):>8.1f}{st.stdev(vals):>7.1f}" if len(vals)>1 else ""
    print(row)
print()
print("=== 人件費率（販管費うち人件費 ＋ 売上原価うち労務費）===")
print(f"{'区分':<12}" + "".join(f"{fy[y]:>8}" for y in years))
print("-"*60)
for k,l in tiers:
    row=f"{l:<12}"
    for y in years:
        g=load(files[(y,3)]); s=g("売上高",k)
        a=(g("販売費及び一般管理費_うち、人件費",k) or 0); b=(g("売上原価_うち、労務費",k) or 0)
        row+=f"{(a+b)/s*100:>8.1f}" if s else f"{'--':>8}"
    print(row)
