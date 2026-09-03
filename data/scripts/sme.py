import openpyxl
def load(fn):
    ws=openpyxl.load_workbook(fn, data_only=True)['＜拡大推計表＞']
    cols={}
    for c in range(9, ws.max_column+1):
        h=ws.cell(6,c).value
        if h: cols[str(h)]=c
    rows={}
    for r in range(7, ws.max_row+1):
        v=ws.cell(r,9).value
        if v: rows[str(v)]=r
    return ws, cols, rows
ws,cols,rows=load('../raw/chusho/r07_03_uriage_hiyo.xlsx')
def g(item, col):
    if item not in rows or col not in cols: return None
    v=ws.cell(rows[item], cols[col]).value
    return v if isinstance(v,(int,float)) else None
tiers=[("合計_法人企業_5人以下","法人 5人以下"),("合計_法人企業_6～20人","法人 6-20人"),
       ("合計_法人企業_21～50人","法人 21-50人"),("合計_法人企業_51人以上","法人 51人以上"),
       ("合計_個人企業","個人企業")]
print("=== 令和6年度 全産業 従業者規模別 ===")
print(f"{'区分':<14}{'企業数':>10}{'売上高':>12}{'粗利率':>8}{'販管費率':>9}{'人件費率':>9}{'営業利益率':>10}")
print("-"*74)
for key,lbl in tiers:
    s=g("売上高",key); n=g("母集団企業数",key)
    if not s: continue
    gp=g("売上総利益",key)/s; sga=g("販売費及び一般管理費",key)/s
    hr=(g("販売費及び一般管理費_うち、人件費",key) or 0)/s
    op=g("営業利益",key)/s
    print(f"{lbl:<14}{n:>10,.0f}{s/1e6:>11,.1f}兆{gp*100:>7.1f}%{sga*100:>8.1f}%{hr*100:>8.1f}%{op*100:>9.1f}%")
print()
print("=== 労務費・減価償却を含めた比較 ===")
print(f"{'区分':<14}{'売原労務費':>11}{'販管人件費':>11}{'人件費計':>10}{'減価償却計':>11}")
print("-"*60)
for key,lbl in tiers:
    s=g("売上高",key)
    if not s: continue
    l1=(g("売上原価_うち、労務費",key) or 0)/s
    l2=(g("販売費及び一般管理費_うち、人件費",key) or 0)/s
    d1=(g("売上原価_うち、減価償却費",key) or 0)/s
    d2=(g("販売費及び一般管理費_うち、減価償却費",key) or 0)/s
    print(f"{lbl:<14}{l1*100:>10.1f}%{l2*100:>10.1f}%{(l1+l2)*100:>9.1f}%{(d1+d2)*100:>10.1f}%")
