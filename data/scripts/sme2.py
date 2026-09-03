import openpyxl
def load(fn):
    ws=openpyxl.load_workbook(fn, data_only=True)['＜拡大推計表＞']
    cols={str(ws.cell(6,c).value):c for c in range(9,ws.max_column+1) if ws.cell(6,c).value}
    rows={str(ws.cell(r,9).value):r for r in range(7,ws.max_row+1) if ws.cell(r,9).value}
    def g(item,col):
        if item not in rows or col not in cols: return None
        v=ws.cell(rows[item],cols[col]).value
        return v if isinstance(v,(int,float)) else None
    return g
gA=load('../raw/chusho/r07_02_shisan_fusai.xlsx')
gS=load('../raw/chusho/r07_03_uriage_hiyo.xlsx')
gI=load('../raw/chusho/r07_06_setsubi_lease.xlsx')
tiers=[("合計_法人企業_5人以下","法人 5人以下"),("合計_法人企業_6～20人","法人 6-20人"),
       ("合計_法人企業_21～50人","法人 21-50人"),("合計_法人企業_51人以上","法人 51人以上")]
print("=== 従業者規模別 κ（法人企業、令和6年度）===")
print(f"{'区分':<14}{'DSO':>7}{'DIO':>7}{'DPO':>7}{'CCC':>8}{'DSO-DPO':>9}")
print("-"*54)
for k,l in tiers:
    s=gS("売上高",k)
    ar=gA("資産_流動資産_うち、受取手形・売掛金",k)
    inv=gA("資産_流動資産_うち、棚卸資産",k)
    ap=gA("負債及び純資産_負債_流動負債_うち、支払手形・買掛金",k)
    if not all([s,ar,inv,ap]): continue
    dso,dio,dpo=ar/s*365, inv/s*365, ap/s*365
    print(f"{l:<14}{dso:>7.0f}{dio:>7.0f}{dpo:>7.0f}{dso+dio-dpo:>8.0f}{dso-dpo:>9.0f}")
print()
print("=== 設備投資（個人企業を含む）===")
print(f"{'区分':<14}{'売上高':>12}{'設備投資額':>12}{'I/r':>8}{'投資実施率':>11}")
print("-"*60)
for k,l in tiers+[("合計_個人企業","個人企業")]:
    s=gS("売上高",k); I=gI("設備投資額",k)
    n=gI("母集団企業数",k); yes=gI("設備投資の状況_設備投資を行った企業数",k)
    if not all([s,I,n,yes]): continue
    print(f"{l:<14}{s/1e6:>11,.1f}兆{I/1e6:>11,.1f}兆{I/s*100:>7.1f}%{yes/n*100:>10.1f}%")
print()
print("=== 資本金階級（法人企業統計）との比較：DSO−DPO ===")
print("  法人企業統計 2024（資本金別）: 10億~ 27日 / 1千万未満 14日")
print("  中小企業実態 2024（従業者別）: 上表のとおり")
