import openpyxl, glob, re, statistics as st
def load(fn):
    wb=openpyxl.load_workbook(fn, data_only=True); ws=wb[wb.sheetnames[0]]
    # 項目名列を自動検出：6行目が「項目名」の列
    ic=None
    for c in range(1,15):
        if str(ws.cell(3,c).value)=='項目名': ic=c; break
    if ic is None:
        for c in range(1,15):
            if str(ws.cell(7,c).value)=='母集団企業数': ic=c; break
    cols={str(ws.cell(6,c).value):c for c in range(ic,ws.max_column+1) if ws.cell(6,c).value}
    rows={str(ws.cell(r,ic).value):r for r in range(7,ws.max_row+1) if ws.cell(r,ic).value}
    def g(item,col):
        if item not in rows or col not in cols: return None
        v=ws.cell(rows[item],cols[col]).value
        return v if isinstance(v,(int,float)) else None
    return g
files={}
for fn in glob.glob('/home/claude/sme/**/*.xlsx', recursive=True):
    m=re.search(r'r(\d\d)_(\d\d)_1_', fn)
    if m: files[(int(m.group(1)), int(m.group(2)))]=fn
fy={2:2019,3:2020,4:2021,5:2022,6:2023,7:2024}
years=sorted({k[0] for k in files})
tiers=[("合計_法人企業_5人以下","法人5人以下"),("合計_法人企業_6～20人","法人6-20"),
       ("合計_法人企業_21～50人","法人21-50"),("合計_法人企業_51人以上","法人51人~"),
       ("合計_個人企業","個人企業")]
def series(tbl, fn_calc):
    out={}
    for k,l in tiers:
        vals=[]
        for y in years:
            key=(y,tbl) if (y,tbl) in files else (y,4 if tbl==6 else tbl)
            if key not in files: vals.append(None); continue
            vals.append(fn_calc(load(files[key]), k, load(files[(y,3)])))
        out[l]=vals
    return out
def show(title, out, fmt="{:>8.1f}"):
    print(f"\n=== {title} ===")
    print(f"{'区分':<12}" + "".join(f"{fy[y]:>8}" for y in years) + f"{'平均':>8}{'SD':>7}")
    print("-"*74)
    for l,vals in out.items():
        v=[x for x in vals if x is not None]
        row=f"{l:<12}" + "".join(fmt.format(x) if x is not None else f"{'--':>8}" for x in vals)
        if len(v)>1: row+=f"{st.mean(v):>8.1f}{st.stdev(v):>7.1f}"
        print(row)
show("営業利益率(%)", series(3, lambda g,k,gs: (g("営業利益",k)/g("売上高",k)*100) if g("売上高",k) else None))
show("人件費率(%)", series(3, lambda g,k,gs: (((g("販売費及び一般管理費_うち、人件費",k) or 0)+(g("売上原価_うち、労務費",k) or 0))/g("売上高",k)*100) if g("売上高",k) else None))
show("減価償却率(%)", series(3, lambda g,k,gs: (((g("販売費及び一般管理費_うち、減価償却費",k) or 0)+(g("売上原価_うち、減価償却費",k) or 0))/g("売上高",k)*100) if g("売上高",k) else None))
show("I/r(%)", series(6, lambda g,k,gs: (g("設備投資額",k)/gs("売上高",k)*100) if (g("設備投資額",k) and gs("売上高",k)) else None))
